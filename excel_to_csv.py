import os
import csv
import openpyxl
import xlrd

def process_xlsx_sheet(sheet, output_path):
	with open(output_path, 'w', newline='', encoding='utf-8') as csv_file:
		csv_writer = csv.writer(csv_file)
		for row in sheet.iter_rows(values_only=True):
			values = prep_row_values(row)
			csv_writer.writerow(values)

def process_xls_sheet(sheet, output_path):
	with open(output_path, 'w', newline='', encoding='utf-8') as csv_file:
		csv_writer = csv.writer(csv_file)
		for row_idx in range(sheet.nrows):
			values = prep_row_values(sheet.row_values(row_idx))
			csv_writer.writerow(values)

def prep_row_values(row):
	return [prep_row_value(x) for x in row]

def prep_row_value(x):
	if isinstance(x, str):
		return x.strip()

	if isinstance(x, float):
		return round(x, 2)
	
	if isinstance(x, int):
		return x
	
	assert x == None, f"unhandled value '{x}'."
	return x


def process_excel_file(filepath, outdir):
	filext = file_ext(filepath)
	filename = os.path.basename(filepath)
	multi_sheet = is_multi_sheet(filepath)
	
	if filext == '.xlsx':
		workbook = openpyxl.load_workbook(filepath, data_only=True)
		if multi_sheet:
			for sheet_name in workbook.sheetnames:
				sheet = workbook[sheet_name]
				output_path = os.path.join(outdir, f"{filename}_{sheet_name}.csv")
				process_xlsx_sheet(sheet, output_path)
		else:
			sheet = workbook.active
			output_path = os.path.join(outdir, f"{filename}.csv")
			process_xlsx_sheet(sheet, output_path)
	
	elif filext == '.xls':
		workbook = xlrd.open_workbook(filepath)
		if multi_sheet:
			for sheet_name in workbook.sheet_names():
				sheet = workbook.sheet_by_name(sheet_name)
				output_path = os.path.join(outdir, f"{filename}_{sheet_name}.csv")
				process_xls_sheet(sheet, output_path)
		else:
			sheet = workbook.sheet_by_index(0)
			output_path = os.path.join(outdir, f"{filename}.csv")
			process_xls_sheet(sheet, output_path)
	
	else:
		raise UnknownFilext(filext)

def is_multi_sheet(filepath):
	filext = file_ext(filepath)
	
	if filext == '.xlsx':
		workbook = openpyxl.load_workbook(filepath, read_only=True)
		return len(workbook.sheetnames) > 1
	elif filext == '.xls':
		workbook = xlrd.open_workbook(filepath, on_demand=True)
		return workbook.nsheets > 1
	else:
		raise UnknownFilext(ext)

def file_ext(filepath):
	_, ext = os.path.splitext(filepath)
	return ext.lower()

def is_tmp_file(filepath):
	filename = os.path.basename(filepath)
	if filename.startswith("~$"):
		return True
	return False

def UnknownFilext(ext):
	return ValueError(f"Unsupported file format: {ext}")


def excel_to_csv(input_dir, output_dir):
	if not os.path.exists(output_dir):
		os.makedirs(output_dir)
	
	for filename in os.listdir(input_dir):
		if filename.endswith(('.xlsx', '.xls')) and not is_tmp_file(filename):
			filepath = os.path.join(input_dir, filename)
			process_excel_file(filepath, output_dir)

def main():
	input_dir = input("Enter the directory path containing Excel files: ")
	output_dir = input("Enter the output directory path for CSV files: ")
	
	process_directory(input_dir, output_dir)
	print(f"Conversion complete. CSV files saved in: {output_dir}")

if __name__ == "__main__":
	main()
