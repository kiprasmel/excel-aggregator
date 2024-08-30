#!/usr/bin/env python3

import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import xlrd
from dataclasses import dataclass
from typing import List, Tuple, Callable, Optional, Any
from datetime import datetime

@dataclass
class Location:
	x: int
	y: int
	value: Any
	sheet: Any
	prefix: str = ""
	suffix: str = ""

	def goRight(self):
		return self._move(1, 0)

	def goBelow(self):
		return self._move(0, 1)

	def goRightUntilExact(self, target):
		return self._moveUntil(1, 0, lambda val: val == target)

	def goRightUntilPrefix(self, prefix):
		return self._moveUntil(1, 0, lambda val: str(val).startswith(prefix))

	def goBelowUntilExact(self, target):
		return self._moveUntil(0, 1, lambda val: val == target)

	def goBelowUntilPrefix(self, prefix):
		return self._moveUntil(0, 1, lambda val: str(val).startswith(prefix))

	def _move(self, dx, dy):
		new_x, new_y = self.x, self.y
		while True:
			new_x += dx
			new_y += dy
			value = self._get_cell_value(new_x, new_y)
			if value is None:  # reached the end of the sheet
				return None
			if value != "":  # found a non-empty cell
				return Location(new_x, new_y, value, self.sheet)

	def _moveUntil(self, dx, dy, condition):
		current = self
		while True:
			current = current._move(dx, dy)
			if not current:
				return None
			if condition(current.value):
				return current

	def _get_cell_value(self, x, y):
		if isinstance(self.sheet, xlrd.sheet.Sheet):
			if x > self.sheet.ncols or y > self.sheet.nrows:
				return None
			return self.sheet.cell_value(y-1, x-1)
		else:  # assume openpyxl sheet
			if x > self.sheet.max_column or y > self.sheet.max_row:
				return None
			return self.sheet.cell(row=y, column=x).value

class Finder:
	def __init__(self, find_func):
		self.find_func = find_func

	def goRight(self):
		return self._chain(lambda loc: loc.goRight())

	def goBelow(self):
		return self._chain(lambda loc: loc.goBelow())

	def goRightUntilExact(self, value):
		return self._chain(lambda loc: loc.goRightUntilExact(value))

	def goRightUntilPrefix(self, prefix):
		return self._chain(lambda loc: loc.goRightUntilPrefix(prefix))

	def goBelowUntilExact(self, value):
		return self._chain(lambda loc: loc.goBelowUntilExact(value))

	def goBelowUntilPrefix(self, prefix):
		return self._chain(lambda loc: loc.goBelowUntilPrefix(prefix))

	def getSuffix(self):
		return self._chain(lambda loc: Location(loc.x, loc.y, loc.suffix, loc.sheet))

	def modify(self, func: Callable[[Any], Any]):
		return self._chain(lambda loc: Location(loc.x, loc.y, func(loc.value), loc.sheet))

	def _chain(self, operation):
		def new_find_func(sheet):
			loc = self.find_func(sheet)
			if loc:
				return operation(loc)
			return None
		return Finder(new_find_func)

	def __call__(self, sheet):
		return self.find_func(sheet)

def findExact(value: str):
	def finder(sheet):
		for y, row in enumerate(sheet.get_rows(), start=1):
			for x, cell in enumerate(row, start=1):
				if cell.value == value:
					return Location(x, y, cell.value, sheet)
		return None
	return Finder(finder)

def findPrefix(prefix: str):
	def finder(sheet):
		for y, row in enumerate(sheet.get_rows(), start=1):
			for x, cell in enumerate(row, start=1):
				if isinstance(cell.value, str) and cell.value.startswith(prefix):
					suffix = cell.value[len(prefix):]
					return Location(x, y, cell.value, sheet, prefix=prefix, suffix=suffix)
		return None
	return Finder(finder)

def aggregate_excel_data(folder_path: str, parse_columns: List[Tuple[str, Callable]]):
	all_data = []
	folder_name = os.path.basename(os.path.normpath(folder_path))
	timestamp = datetime.now().strftime("%Y-%m-%d_%H:%M:%S")
	outdir = "out"
	Path(outdir).mkdir(parents=True, exist_ok=True)
	output_file_excel = os.path.join(outdir, f"aggregated-{folder_name}-{timestamp}.xlsx")
	# output_file_csv = f"aggregated-{folder_name}-{timestamp}.csv"

	for filename in os.listdir(folder_path):
		if filename.endswith('.xlsx') or filename.endswith('.xls'):
			file_path = os.path.join(folder_path, filename)
			
			if filename.endswith('.xlsx'):
				workbook = load_workbook(filename=file_path, data_only=True)
				sheet = workbook.active
			else:  # .xls file
				workbook = xlrd.open_workbook(file_path)
				sheet = workbook.sheet_by_index(0)

			row_data = {'Filename': filename}
			for column_name, value_fn in parse_columns:
				location = value_fn(sheet)
				if location:
					row_data[column_name] = location.value

			all_data.append(row_data)

	df = pd.DataFrame(all_data)
	df.to_excel(output_file_excel, index=False)
	print(f"Data aggregated and saved to {output_file_excel}")
	# df.to_csv(output_file_csv, index=False)
	# print(f"Data aggregated and saved to {output_file_csv}")

def main():
	parse_columns = [
		("SF NUMERIS", findPrefix("PVM SĄSKAITA FAKTŪRA (VAT INVOICE)").modify(lambda x: x.split(" ")[-1])),
		("DATA", findPrefix("Išrašymo data / Date:").getSuffix()),
		("KODAS", findPrefix("Pirkėjas / Buyer").goBelowUntilPrefix("Įmones kodas:").getSuffix()),
		("PVM KODAS", findPrefix("Pirkėjas / Buyer").goBelowUntilPrefix("PVM kodas:").getSuffix()),
		("VARDAS PAVARDĖ/ĮM. PAVADINIMAS", findExact("Pirkėjas / Buyer").goBelow()),
		("KAINA BE PVM", findExact("Bendros sumos EUR").goBelowUntilExact("Suma be PVM / total amount:").goRight()),
	]

	folder_path = input("Enter the folder path containing Excel files: ")
	print("")
	aggregate_excel_data(folder_path, parse_columns)

if __name__ == "__main__":
	main()
